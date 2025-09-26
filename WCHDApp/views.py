from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from .models import Fund, Testing, Item, Grant, GrantLine, Revenue, Expense, Line
from django.db.models.fields.related import ForeignKey, ManyToManyField, OneToOneField
from .forms import FundForm, TableSelect, InputSelect, ExportSelect,reconcileForm, activitySelect, FileInput
from django.forms import modelform_factory, Select
from django import forms
from django.apps import apps
from django.db.models import DecimalField, AutoField
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import permission_required
from django.contrib import messages
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
from datetime import datetime
import json
from django.shortcuts import render
from django.urls import reverse
from django.utils.timezone import now
from django.utils.dateparse import parse_datetime
from django.core.exceptions import ValidationError
import re

def generate_pdf(request, tableName):
    buffer = BytesIO()

    # Create a PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()


    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=11,
        fontName="Helvetica-Bold"
    )

    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.black,
        spaceAfter=6,
        fontName="Helvetica-Oblique"
    )


    #logo = Image("logo.png", width=80, height=80)
    #logo.hAlign = 'LEFT'
    #elements.append(logo)

    elements.append(Spacer(1, 12))

    # Report Title
    elements.append(Paragraph("Washington County Health Department", title_style))
    elements.append(Paragraph(
        "List of Active Grants. Active means they have been awarded and the final expenditure report has not yet been approved.",
        subtitle_style
    ))

    elements.append(Spacer(1, 12))

    #Same logic as tableView, needs updated to current 
    model = apps.get_model('WCHDApp', tableName)
    values = model.objects.all().values()
    fields = model._meta.get_fields()
    fieldNames = []
    decimalFields = []
    aliasNames = []
    for field in fields:
        if field.is_relation:
            if field.auto_created:
                continue
            else:
                parentModel = apps.get_model('WCHDApp', field.name)
                fkName = parentModel._meta.pk.name
                fkAlias = parentModel._meta.pk.verbose_name
                aliasNames.append(fkAlias)
                fieldNames.append(fkName)

        else:
            if isinstance(field, DecimalField):
                decimalFields.append(field.name)
            aliasNames.append(field.verbose_name)  
            fieldNames.append(field.name)

    data = [
        aliasNames,
    ]
    
    for row in values:
        print(row)
        line = []
        for field in fieldNames:
            line.append(row[field])
        data.append(line)


    # Table Styling
    table = Table(data, colWidths=[80, 70, 70, 70, 70, 50, 100, 50, 90, 40])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkgray),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
    ]))

    elements.append(table)

    # Totals (below table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>Total Active Grants:</b> $571,880.00", styles["Normal"]))
    elements.append(Paragraph("<b>Total Amount for Project:</b> $19,144.00", styles["Normal"]))

    #Build PDF
    doc.build(elements)

    # Get the PDF value from buffer
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="testing.pdf"'

    return response

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def reconcile(request):
    if request.method == "POST":
        form = reconcileForm(request.POST, request.FILES)
        if form.is_valid():
            firstFile = form.cleaned_data['firstFile'] 
            secondFile = form.cleaned_data['secondFile'] 
            
            df1 = pd.read_csv(firstFile)
            df2 = pd.read_csv(secondFile)

            columnList = list(df1.columns)
            for i in range(len(columnList)):
                columnList[i] = columnList[i].strip()
            # Merge the two lists and remove duplicates
            merged_df = pd.concat([df1, df2]).drop_duplicates()

            # Identify common entries (entries in both list1 and list2)
            common_entries = df1.merge(df2, on=columnList, how="inner")

            # Save merged data to an Excel file
            output_file = "output.xlsx"
            merged_df.to_excel(output_file, index=False)

            # Load the saved Excel file for formatting
            wb = load_workbook(output_file)
            ws = wb.active

            # Define highlight style
            highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Convert common entries into a set for fast lookup
            rows = common_entries[columnList].apply(tuple, axis=1)
            common_set = set(rows)
            print(common_set)

            # Apply highlighting to rows that are NOT common (i.e., unique to either list1 or list2)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1):
                rowData = []
                for column in row:
                    rowData.append(column.value)
                rowTuple = tuple(rowData)
                print(rowTuple)
                if rowTuple not in common_set:  # Highlight unique entries only
                    for cell in row:
                        cell.fill = highlight_fill
            
            #Saveing to stream in file attachment format
            outputStream = BytesIO()
            wb.save(outputStream)
            outputStream.seek(0)
            response = HttpResponse(outputStream.getvalue(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="reconciliation.xlsx"'

            return response
    else:
        form = reconcileForm()
    return render(request, "WCHDApp/reconcile.html", {"form":form})

#This view is used to select what table we want to create a report from
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def reports(request):
    if request.method == "POST":
        #TableSelect is a form defined in forms.py
        form = TableSelect(request.POST)

        #Take the data from the form and pass it to our pdf generator function
        button = request.POST.get('button')
        if form.is_valid():
            tableName = form.cleaned_data['table'] 
            if button == "daily":
                return redirect('dailyReport')
            else:
                return redirect('generate_pdf', tableName)
    else:
        form = TableSelect()
    return render(request, "WCHDApp/reports.html", {'form': form})

def index(request):
    # Set session start time if it's not already set
    if not request.session.get('session_start_time'):
        request.session['session_start_time'] = str(now())

    # Calculate session duration
    session_start_str = request.session.get('session_start_time')
    duration_display = "0h 0m 0s"  # Default

    todayDate = datetime.today()

    expenses = Expense.objects.filter(date=todayDate)
    revenues = Revenue.objects.filter(date=todayDate)

    expenseTotal = 0
    revenueTotal = 0
    for expense in expenses:
        expenseTotal += expense.amount

    for revenue in revenues:
        revenueTotal += revenue.amount

    if session_start_str:
        session_start = parse_datetime(session_start_str)
        if session_start:
            duration = now() - session_start
            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            duration_display = f"{hours}h {minutes}m {seconds}s"
    
    context ={
        'duration': duration_display,
        "revenueTotal": revenueTotal,
        "expenseTotal": expenseTotal
    }

    # Pass formatted string to template
    return render(request, "WCHDApp/index.html", context)

#Login page logic
def logIn(request):
    #Getting username and password from the form
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        #If authentication is successful, returns related user object
        #notAdmin pass is Marietta123
        #Django implemented function to check databse for user and rights etc
        user = authenticate(request, username = username, password=password)

        #If authenticated, login (Django function) and redirect to hub
        #If not give error 
        if user is not None:
            login(request,user)
            return redirect('index')
        else:
            messages.error(request, "Invalid username or password")
    return render(request, "WCHDApp/logIn.html")

#Logic to get what tables we want to see/create from. Same thing as reports
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def viewTableSelect(request):
    if request.method == 'POST':
        form = TableSelect(request.POST)

        #Logic to figure out which button sent the request so that we can correctly redirect
        #Each button has a different value aligning to their names, this is set in the html file "viewTableSelect"
        button = request.POST.get('button')
        if form.is_valid():
            tableName = form.cleaned_data['table'] 
            if tableName == 'Payroll':
                return redirect('payrollView')
            elif tableName == "Transaction":
                return redirect('transactionCustomView')
            elif tableName == "GrantExpense":
                return redirect('grantExpenses')
            elif tableName == "Expense":
                return redirect('transactionsExpenses')
            elif tableName == "Revenue":
                return redirect('transactionsItem')
            elif tableName == "Line":
                return redirect('lineView')
            elif tableName == "GrantLine":
                return redirect('grantLineView')
            if button == "seeTable":
                return redirect('tableView', tableName)
            elif button == "create":
                return redirect('createEntry', tableName)
    else:
        form = TableSelect()
    return render(request, "WCHDApp/viewTableSelect.html", {'form': form})

#This function decides what data we use in our tables in tableView.html
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def tableView(request, tableName):

    #Grabbing the model selected in viewTableSelect
    model = apps.get_model('WCHDApp', tableName)

    #Getting data from that model
    values = model.objects.all()

    #Getting just field names from model
    #Use .fields instead of .get_fields() because we do not want reverse relationships
    fields = model._meta.fields

    #Any property that we define in models need to go here so our logic can include them in the table
    calculatedProperties = {
        "Testing": [("fundBalanceMinus3", "Fund Balance Minus 3")],
        "Benefits": [("pers", "Public Employee Retirement System"), ("medicare", "Medicare"),("wc", "Workers Comp"), ("plar", "Paid Leave Accumulation Rate"), ("vacation", "Vacation"), ("sick", "Sick Leave"), ("holiday", "Holiday Leave"), ("total_hrly", "Total Hourly Cost"), ("percent_leave", "Percent Leave"), ("monthly_hours", "Monthly Hours"), ("board_share_hrly", "Board Share Hourly"), ("life_hourly", "Life Hourly"), ("salary", "Salary"), ("fringes", "Fringes"), ("total_comp", "Total Compensation")],
        "Payroll": [("pay_rate", "Pay Rate")],
        "Fund":[("calcRemaining", "Remaining"), ("budgeted", "Budgeted")],
        "GrantLine": [("budgetRemaining", "Budget Remaining"), ("budgetSpent", "Budget Spent"), ("totalIncome", "Total Income")],
        "Grant": [("grantAwardAmountRemaining", "Grant Award Amount Remaining")]
    }

    #This is used to decide which fields we want to show in the accumulator based on each model
    summedFields = {
        "Fund": "fund_cash_balance", 
        "Line": "line_total_income",
        "Transaction": "amount",
    }
    

    fieldNames = []
    aliasNames = []
    decimalFields = []
    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)
    #Making sure properties are added like normal fields to the tables
    if tableName in calculatedProperties:
        for property in calculatedProperties[tableName]:
            #print(property)
            aliasNames.append(property[1])
            fieldNames.append(property[0])
            decimalFields.append(property[0])

    #Getting values based on if we defined them in summedFields in order to make accumulator
    if tableName in summedFields:
        field = summedFields[tableName]
        accumulator = 0
        for value in values:
            accumulator += getattr(value, field)
        context = {"fields": fieldNames, "aliasNames": aliasNames, "data": values, "tableName": tableName, "decimalFields": decimalFields, "accumulator": accumulator}
    else:
        context = {"fields": fieldNames, "aliasNames": aliasNames, "data": values, "tableName": tableName, "decimalFields": decimalFields}

    return render(request, "WCHDApp/tableView.html", context)

#New system to dynamically create forms based of model
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def createEntry(request, tableName):
    message = ""
    #Grabbing selected model in viewTableSelect
    model = apps.get_model('WCHDApp', tableName)

    if request.method == 'POST':
        #Django function that makes a form based off a provided model
        #FORM UPDATES IF NEEDED, MAKE SURE TO ADD EXCLUSIONS IN NON_POST RENDER AS WELL
        if tableName == "GrantLine":
            form = modelform_factory(model, exclude=["line_budget_spent", "line_budget_remaining"])(request.POST)
        elif tableName == "Fund":
            form = modelform_factory(model, exclude=["fund_total", "fund_budgeted", "fund_remaining"])(request.POST)
        elif tableName == "Line":
            form = modelform_factory(model, exclude=["line_budget_spent", "line_budget_remaining"])(request.POST)
        else:
            form = modelform_factory(model, fields="__all__")(request.POST)

        #Data validation then save to table linked to the model
        #FORM VALIDATION IF NEEDED
        if tableName == "GrantLine":
            if form.is_valid():
                line = form.save(commit=False)

                #Getting input from form
                budgetedAmount = float(request.POST["line_budgeted"])

                #Selected Grant from previous screen
                grantID = request.POST['grant']
                grantModel = apps.get_model("WCHDApp", "Grant")
                grant = grantModel.objects.get(pk=grantID)


                grantLineModel = apps.get_model("WCHDApp", "GrantLine")
                grantLines = grantLineModel.objects.filter(grant=grant)

                #Getting total money that is previously budgeted to lines
                total = 0
                hasReceivedLine = False
                for lineIterable in grantLines:
                    total += lineIterable.line_budgeted
                    if lineIterable.receivingLine == True:
                        hasReceivedLine = True
                grantAwardAmount = grant.award_amount
                grantAwardAmountRemaining = grantAwardAmount - total

                if grantAwardAmountRemaining >= budgetedAmount:
                    line.line_budget_spent = 0
                    line.line_budget_remaining = budgetedAmount
                    if (hasReceivedLine == True) and (line.receivingLine == True):
                        message = "Already has a specified line to receive reimbursement"
                    else:
                        line.save()
                else:
                    message = "Budgeted is more than is left in Grant Award"
        elif tableName == "Fund":
            if form.is_valid():
                fund = form.save(commit=False)
                balance = fund.fund_cash_balance
                baseID = fund.fund_id
                fund.fund_total = balance
                fund.fund_budgeted = 0
                fund.fund_remaining = balance
                currentDateTime = datetime.now()
                year = currentDateTime.year
                fullID = f"{year}-{baseID}"
                fund.fund_id = fullID

                form.save()
                return redirect('tableView', tableName)
        elif tableName == "Line":
            if form.is_valid():
                line = form.save(commit=False)
                budgeted = line.line_budgeted
                line.line_budget_spent = 0
                line.line_budget_remaining = budgeted

                fund = line.fund
                remaining = fund.fund_total - fund.fund_budgeted
                if (remaining >= budgeted):
                    fund.fund_budgeted += budgeted
                    fund.save()
                    line.save()
                    return redirect('tableView', tableName)
                else:
                    message="Not enough remaining balance in fund"
        else:
            if form.is_valid():
                form.save()
                return redirect('tableView', tableName)
            else:
                print(form.errors)
    else:
        if tableName == "GrantLine":
            form = modelform_factory(model, exclude=["line_budget_spent", "line_budget_remaining"])(request.POST)
        elif tableName == "Fund":
            form = modelform_factory(model, exclude=["fund_total", "fund_budgeted", "fund_remaining"])(request.POST)
        elif tableName == "Line":
            form = modelform_factory(model, exclude=["line_budget_spent", "line_budget_remaining"])(request.POST)
        else:
            form = modelform_factory(model, fields="__all__")(request.POST)
    return render(request, "WCHDApp/createEntry.html", {"form": form, "tableName": tableName, "message": message})

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def imports(request):
    message = ""
    if request.method == 'POST':
        form = InputSelect(request.POST, request.FILES)
        if form.is_valid():
            #Pulls data from submitted files
            tableName = form.cleaned_data['table']
            if tableName == 'Payroll':
                return redirect('clockifyImportPayroll')
            selectedFile = form.cleaned_data['file']
            file = pd.read_csv(selectedFile)
            columns = file.columns
            row = file.iloc[0]
            data = []

            #Grab slected model
            model = apps.get_model('WCHDApp', tableName)
            fields = model._meta.get_fields()

            #Define what fields we want to look at from the file and model
            neededFields = []
            for field in fields:
                if field.is_relation:
                        #fks.append(field.name)
                        if field.auto_created:
                            continue
                        else:
                            #Grab related model. This is why foreign keys have to be named after the model 
                            parentModel = apps.get_model('WCHDApp', field.name)

                            #Get the related models primary key
                            fkName = parentModel._meta.pk.name
                            neededFields.append(fkName)
                else:
                    neededFields.append(field.name)
            if neededFields != list(columns):
                message = "Bad File. Please check your CSV format and try again."
                return render(request, "WCHDApp/imports.html", {"form": form, "message": message})
            lookUpFields = []
            fks = []
            #Same logic as above just for lookups and fk
            for field in fields:
                #Logic for foreign keys
                if field.is_relation:
                    fks.append(field.name)
                    if field.auto_created:
                        continue
                    else:
                        #Grab related model. This is why foreign keys have to be named after the model 
                        parentModel = apps.get_model('WCHDApp', field.name)

                        #Get the related models primary key
                        fkName = parentModel._meta.pk.name
                        #fks.append(fkName)
                        #Primary keys verbose name
                        fkAlias = parentModel._meta.pk.verbose_name
                else:
                    lookUpFields.append(field)
            
            #Creating a dictionary for each row in the file
            for i in range(len(file)):
                dict = {}
                row = file.iloc[i]
                for column in columns:
                    dict[column] = row[column]
                data.append(dict)
                
            #For each entry in the dictionary convert types and then create an object based on the dict
            for line in data:
                for key in line:
                    if type(line[key]) == np.int64:
                        line[key] = int(line[key])
                    if key in fks:
                        parentModel = apps.get_model('WCHDApp', key)
                        print("Grab the object linked")
                        line[key] = parentModel.objects.get(pk=line[key])
                print(line)
                obj, _ = model.objects.update_or_create(
                    **line,
                    defaults = line
                )    
    else:
        form = InputSelect()
    
        
    return render(request, "WCHDApp/imports.html", {"form": form, "message": message})

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def exports(request):

    message = ""
    if request.method == 'POST':
        form = ExportSelect(request.POST)
        if form.is_valid():
            tableName = form.cleaned_data['table']
            fileName = form.cleaned_data['fileName']
            
            model = apps.get_model('WCHDApp', tableName)
            data = model.objects.all().values()
            exportData = pd.DataFrame.from_records(data)

            #From what I read the 2 commented lines are how we can show it in a new tab before download
            #However, its raw text apparently browsers dont like not immediately downloading csv, could be useful for our reports though
            response = HttpResponse(content_type='text/csv')
            #response = HttpResponse(content_type='text/text')
            #response['Content-Disposition'] = f'inline; filename="{fileName}.csv"'
            response['Content-Disposition'] = f'attachment; filename="{fileName}.csv"'

            exportData.to_csv(path_or_buf=response, index=False)
            return response
    else:
        form = ExportSelect()
        
    return render(request, "WCHDApp/exports.html", {"form": form, "message": message})

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def countyPayrollExport(request):
    payperiodModel = apps.get_model('WCHDApp', "PayPeriod")
    payperiods = payperiodModel.objects.all()

    if request.method == "POST":
        payperiod = request.POST.get('payPeriod')
        fileName = request.POST.get('fileName')

        payrollModel = apps.get_model('WCHDApp', "Payroll")
        entries = payrollModel.objects.select_related('employee', "ActivityList").filter(payperiod__payperiod_id = payperiod)

        employeeHoursByActivity = {}
        codeMappings = {
            "SICK": "S",
            "COMP": "C",
            "VAC": "V",
            "HOLIDAY": "H"
        }

        secondaryMapping = {
            "S": "S-SICK",
            "C":  "C-COMPTIME",
            "V": "V-VACATION",
            "H": "H-HOLIDAY",
            "R": "R-REGULAR PA"
        }

        for entry in entries:
            activityName = entry.ActivityList.program.upper()
            for keyword, code in codeMappings.items():
                if keyword in activityName:
                    paycode = code
                    break
                else:
                    paycode = "R"
            paycodeName = secondaryMapping[paycode]

            if paycodeName not in employeeHoursByActivity:
                employeeHoursByActivity[paycodeName] = {}

            if entry.employee in employeeHoursByActivity[paycodeName]:
                employeeHoursByActivity[paycodeName][entry.employee] += entry.hours
            else:
                employeeHoursByActivity[paycodeName][entry.employee] = entry.hours

        exportData = []
        for activity, employeeDict in employeeHoursByActivity.items():
            for employee, hours in employeeDict.items():
                exportData.append({
                    "JobNumber": employee.employee_id,
                    "Paycode": activity,
                    "Time Group/Description": "",
                    "Hours": hours,
                    "HourlyRate": employee.pay_rate,
                    "Salary": "",
                    "AccountDistribution": employee.gen_pay_fund.fund_id
                })
        exportData = pd.DataFrame(exportData)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{fileName}.csv"'
        exportData.to_csv(path_or_buf=response, index=False)

        return response
        """
        for entry in entries:
            if entry.ActivityList not in employeeHoursByActivity:
                if entry.ActivityList == 
                employeeHoursByActivity[entry.ActivityList] = {}
            if entry.employee in employeeHoursByActivity[entry.ActivityList]:
                employeeHoursByActivity[entry.ActivityList][entry.employee] += entry.hours
            else:
                employeeHoursByActivity[entry.ActivityList][entry.employee] = entry.hours
        """
        """
        Setting up dictionary to track workers total hours
        employeeHoursByActivity = {
            Gen Pay {
                empObject:80,
                empObject:90
            }
            Sick {
                empObject: 70
            }
        }
        """

        """
        employeeHours = {}
        for entry in entries:
            if entry.employee in employeeHours:
                employeeHours[entry.employee] += entry.hours
            else:
                employeeHours[entry.employee] = entry.hours

        exportData = []
        for employee, hours in employeeHours.items():
            #Gotta figure out how to do paycode
            exportData.append({
                "JobNumber": employee.employee_id,
                "Paycode": "R-REGULAR PA",
                "Time Group/Description": "",
                "Hours": hours,
                "HourlyRate": employee.pay_rate,
                "Salary": "",
                "AccountDistribution": employee.gen_pay_fund.fund_id
            })
        exportData = pd.DataFrame(exportData)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{fileName}.csv"'
        exportData.to_csv(path_or_buf=response, index=False)

        return response
        print(exportData)
        """

    context = {
        "payperiods": payperiods
    }
    return render(request, "WCHDApp/countyPayrollExport.html", context)

#This is for revenue but was named previous to table split
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def transactionsItem(request):
    itemModel = apps.get_model('WCHDApp', "Item")
    itemValues = itemModel.objects.filter(line__lineType="Revenue")
    if request.method == "POST":
        itemID = request.POST.get('itemSelect')
        return redirect(transactionsView,itemID)
    
    return render(request, "WCHDApp/transactionsItem.html", {"items":itemValues})

def transactionsView(request):
    message = ""
    revenueModel = apps.get_model('WCHDApp', "revenue")
    itemID = request.GET.get('itemSelect')
    revenueValues = revenueModel.objects.filter(item_id=itemID)

    #Getting just field names from model
    fields = revenueModel._meta.get_fields()

    #Lists to sort fields for styling
    fieldNames = []
    decimalFields = []
    aliasNames = []

    #Fields that should be accumulated
    summedFields = {
        "Fund": "fund_cash_balance", 
        "Line": "line_total_income",
        "Transaction": "amount",
    }

    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)

            
    #Making the view for the cashiers to be able to see and add transaction on the same page
    RevenueForm = modelform_factory(revenueModel, exclude=(["item", "date", "line", "employee"]),  
                                    widgets={
                                        'people': forms.Select(attrs={'class': 'searchable-select'}),
                                        'grantLine': forms.Select(attrs={'class': 'searchable-select'}),
                                    })

    #Getting values from our db so they dont have to
    item = Item.objects.get(pk=itemID)

    if request.method == 'POST':
        form = RevenueForm(request.POST)
        if form.is_valid():
            #Create the instance but don't save it yet
            revenue = form.save(commit=False)
            user  = request.user
            employeeGrab = False
            try:
                employeeModel = apps.get_model('WCHDApp', "employee")
                employee = employeeModel.objects.get(user=user)
                revenue.employee = employee
                employeeGrab = True
            except:
                message = "No employee with signed in user"
            #Adding the values from before
            #transaction.fund = fund
            #transaction.line = line
            revenue.item = item
            revenue.line = item.line
            line = item.line
            if employeeGrab:
                if line.lineType == "Revenue":
                    grantLine = revenue.grantLine
                    if grantLine:
                        grantLine.line_total_income += revenue.amount
                        if grantLine.lineType == "Revenue":
                            grant = grantLine.grant
                            grant.received += revenue.amount
                            grant.save()
                        grantLine.save()
                    

                    line.line_total_income += revenue.amount
                    revenue.save()

                    fund = item.fund
                    fund.fund_cash_balance += revenue.amount
                    fund.save()
                    message = "Revenue Posted Successfully"
                    form = RevenueForm()
                else:
                    message = "Please select a revenue line"

    else:
        form = RevenueForm()

    #return render(request, "WCHDApp/transactionsView.html", {"item": itemID, "revenue": revenueValues,"fields": fieldNames, "aliasNames": aliasNames, "data": revenueValues, "decimalFields": decimalFields, "form":form})
    return render(request, "WCHDApp/partials/revenueTableAndForm.html", {"itemObj":item, "item": itemID, "revenue": revenueValues,"fields": fieldNames, "aliasNames": aliasNames, "data": revenueValues, "decimalFields": decimalFields, "form":form, "message":message})


def addPeopleForm(request):
    peopleModel = apps.get_model("WCHDApp", "people")
    PeopleForm = modelform_factory(peopleModel, fields="__all__")
    itemID = request.GET.get("itemID")
    source = request.GET.get("source")

    if request.method == "POST":
        form = PeopleForm(request.POST)
        itemID = request.POST.get("itemID")
        source = request.POST.get("source")
        if form.is_valid():
            form.save()
            if source == "revenue":
                return redirect('transactionsItem')
            else:
                return redirect('transactionsExpenses')
    else:
        form = PeopleForm()
    context ={
        "form": form,
        "itemID": itemID,
        "source": source
    }

    return render(request, "WCHDApp/partials/formPartial.html", context)

#these are named transaction expense because it was originally built on the transactions table whihc then got split into 2 different tables
@permission_required('WCHDApp.has_full_access', raise_exception=True)
def transactionsExpenses(request):
    itemModel = apps.get_model("WCHDApp", "Item")
    items = itemModel.objects.filter(line__lineType="Expense")


    if request.method=="POST":
        print("Submitted")
    context = {
        "items": items
    }
    return render(request, "WCHDApp/transactionsExpenses.html", context)

def transactionsExpenseTableUpdate(request):
    message = ""
    itemID = request.GET.get('item')
    #print(itemID)
    expenseModel = apps.get_model('WCHDApp', "expense")
    expenseValues = expenseModel.objects.filter(item_id=itemID)

    #Getting just field names from model
    fields = expenseModel._meta.get_fields()

    #Lists to sort fields for styling
    fieldNames = []
    decimalFields = []
    aliasNames = []

    #Fields that should be accumulated
    summedFields = {
        "Fund": "fund_cash_balance", 
        "Line": "line_total_income",
        "Transaction": "amount",
    }

    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)

    #Making the view for the cashiers to be able to see and add transaction on the same page
    expenseForm = modelform_factory(expenseModel, exclude=(["item", "date", "line", "employee"]),  
                                    widgets={
                                        'people': forms.Select(attrs={'class': 'searchable-select'}),
                                        'grantLine': forms.Select(attrs={'class': 'searchable-select'}),
                                    })

    #Getting values from our db so they dont have to
    item = Item.objects.get(pk=itemID)  
    #print(item)

    if request.method == 'POST':
        #print("SUBMITTED ON TABLE UPDATE")
        form = expenseForm(request.POST)
        form.instance.item = item
        #print(request.POST)
        if form.is_valid():
            #Create the instance but don't save it yet
            expense = form.save(commit=False)
            user  = request.user
            employeeGrab = False
            try:
                employeeModel = apps.get_model('WCHDApp', "employee")
                employee = employeeModel.objects.get(user=user)
                expense.employee = employee
                employeeGrab = True
            except:
                message = "No employee with signed in user"
            #Adding the values from before
            #transaction.fund = fund
            #transaction.line = line
            expense.item = item
            expense.line = item.line
            
            fund = item.fund
            line = item.line
            if employeeGrab:
                if line.lineType == "Expense":
                    if expense.grantLine:
                        grantLine = expense.grantLine
                        if (fund.fund_cash_balance >= expense.amount) and (grantLine.line_budgeted >= expense.amount) and (line.line_budget_remaining >= expense.amount):
                            fund.fund_cash_balance -= expense.amount
                            grantLine.line_budget_remaining -= expense.amount
                            grantLine.line_budget_spent += expense.amount
                            line.line_budget_remaining -= expense.amount
                            line.line_budget_spent += expense.amount
                            line.save()
                            grantLine.save()
                            expense.save()
                            fund.save()
                            message = "Expense Posted"
                        else:
                            if (line.line_budget_remaining < expense.amount):
                                message = "Not Enough Line Budgeted"
                            elif (grantLine.line_budgeted < expense.amount):
                                message = "Not Enough Grant Line Budgeted"
                            else:
                                message = "Not enough Fund Cash Balance"
                            
                    else:
                        if (fund.fund_cash_balance >= expense.amount) and (line.line_budget_remaining >= expense.amount):
                            fund.fund_cash_balance -= expense.amount
                            line.line_budget_remaining -= expense.amount
                            line.line_budget_spent += expense.amount
                            line.save()
                            expense.save()
                            fund.save()
                            message = "Expense Posted Successfully"
                            form = expenseForm()
                        else:
                            if (line.line_budget_remaining < expense.amount):
                                message = "Not Enough Line Budgeted"
                            else:
                                message = "Not enough Fund Cash Balance"
                else:
                    message = "Please select an expense line"
            
    else:
        form = expenseForm()

    line = item.line
    context = {
        "expenses": expenseValues,
        "fields": fieldNames, 
        "aliasNames": aliasNames, 
        "data": expenseValues, 
        "decimalFields": decimalFields,
        "form": form,
        "item": item,
        "message": message,
        "budgeted_remaining": line.budgetRemaining,
    }

    return render(request, "WCHDApp/partials/transactionsTablePartial.html", context)

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def lineView(request):
    funds = Fund.objects.all()
    
    context = {
        "funds": funds
    }
    return render(request, "WCHDApp/lineView.html", context)

def lineTableUpdate(request):
    message = ""
    fundID = request.GET.get("fund")
    fund = Fund.objects.get(pk=fundID)
    
    Line = apps.get_model('WCHDApp', "line")
    lines = Line.objects.filter(fund=fund)

    #Getting just field names from model
    fields = Line._meta.fields

    #Lists to sort fields for styling
    fieldNames = []
    decimalFields = []
    aliasNames = []

    calculatedProperties = {
        "Testing": [("fundBalanceMinus3", "Fund Balance Minus 3")],
        "Benefits": [("pers", "Public Employee Retirement System"), ("medicare", "Medicare"),("wc", "Workers Comp"), ("plar", "Paid Leave Accumulation Rate"), ("vacation", "Vacation"), ("sick", "Sick Leave"), ("holiday", "Holiday Leave"), ("total_hrly", "Total Hourly Cost"), ("percent_leave", "Percent Leave"), ("monthly_hours", "Monthly Hours"), ("board_share_hrly", "Board Share Hourly"), ("life_hourly", "Life Hourly"), ("salary", "Salary"), ("fringes", "Fringes"), ("total_comp", "Total Compensation")],
        "Payroll": [("pay_rate", "Pay Rate")],
        "Fund":[("calcRemaining", "Remaining"), ("budgeted", "Budgeted")],
        "Line": [("budgetRemaining", "Budget Remaining"), ("budgetSpent", "Budget Spent"), ("totalIncome", "Total Income")]
    }

    #Fields that should be accumulated
    summedFields = {
        "Fund": "fund_cash_balance", 
        "Line": "line_total_income",
        "Transaction": "amount",
    }

    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)
    #Making sure properties are added like normal fields to the tables
    if "Line" in calculatedProperties:
        for property in calculatedProperties["Line"]:
            #print(property)
            aliasNames.append(property[1])
            fieldNames.append(property[0])
            decimalFields.append(property[0])
    if request.method == 'POST':
        form = modelform_factory(Line, exclude=["fund"])(request.POST)
        form.instance.fund = fund
        if form.is_valid():
            line = form.save()
            message = "Line created successfully"
            form = modelform_factory(Line, exclude=["fund"])()
        else:
            errors = form.errors
            if errors.get("line_budgeted"):
                message = errors["line_budgeted"][0]     
 
    else:
        form = modelform_factory(Line, exclude=["fund"])()
    

    #remainingToBudget = fund.fund_cash_balance - fund.fund_budgeted
    context = {
        "fields": fieldNames, 
        "aliasNames": aliasNames, 
        "data": lines, 
        "decimalFields": decimalFields,
        "form": form,
        "fund": fund,
        "message": message,
        "remainingToBudget": fund.remainingToBudget
    }

    return render(request, "WCHDApp/partials/lineTableUpdate.html", context)

def dailyReport(request):
    buffer = BytesIO()

    # Create a PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()


    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=11,
        fontName="Helvetica-Bold"
    )

    subtitle_style = ParagraphStyle(
        "SubtitleStyle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=colors.black,
        spaceAfter=6,
        fontName="Helvetica-Oblique"
    )


    #logo = Image("logo.png", width=80, height=80)
    #logo.hAlign = 'LEFT'
    #elements.append(logo)

    elements.append(Spacer(1, 12))

    # Report Title
    elements.append(Paragraph("Washington County Health Department", title_style))
    elements.append(Paragraph(
        "List of Active Grants. Active means they have been awarded and the final expenditure report has not yet been approved.",
        subtitle_style
    ))

    elements.append(Spacer(1, 12))

    #Same logic as tableView, needs updated to current 
    model = apps.get_model('WCHDApp', 'transaction')
    today = datetime.today().strftime('%Y-%m-%d')
    values = model.objects.filter(date=today).values()
    fields = model._meta.get_fields()
    fieldNames = []
    decimalFields = []
    aliasNames = []
    for field in fields:
        if field.is_relation:
            if field.auto_created:
                continue
            else:
                parentModel = apps.get_model('WCHDApp', field.name)
                fkName = parentModel._meta.pk.name
                fkAlias = parentModel._meta.pk.verbose_name
                aliasNames.append(fkAlias)
                fieldNames.append(fkName)

        else:
            if isinstance(field, DecimalField):
                decimalFields.append(field.name)
            aliasNames.append(field.verbose_name)  
            fieldNames.append(field.name)

    data = [
        aliasNames,
    ]
    
    for row in values:
        print(row)
        line = []
        for field in fieldNames:
            line.append(row[field])
        data.append(line)


    # Table Styling
    table = Table(data, colWidths=[80, 70, 70, 70, 70, 50, 100, 50, 90, 40])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkgray),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
    ]))

    elements.append(table)

    # Totals (below table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("<b>Total Active Grants:</b> $571,880.00", styles["Normal"]))
    elements.append(Paragraph("<b>Total Amount for Project:</b> $19,144.00", styles["Normal"]))

    #Build PDF
    doc.build(elements)

    # Get the PDF value from buffer
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="testing.pdf"'

    return response

def testing(request):
    model = apps.get_model("WCHDApp", "Employee")
    objects = model.objects.all()
    fields = model._meta.fields
    fieldNames = []
    verboseNames = []
    for field in fields:
        verboseNames.append(field.verbose_name)
        fieldNames.append(field.name)
    
    context = {
        "objects": objects,
        "verboseNames": verboseNames,
        "fieldNames": fieldNames
    }

    return render(request, "WCHDApp/testing.html", context)

def checkPrivileges(request):
    print("Checking privileges")
    if (request.user.is_staff):
        print("Staff")
        return redirect(noPrivileges)  
    else:
        return None 
    
def noPrivileges(request, exception):
    return render(request, "WCHDApp/noPrivileges.html")

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def clockifyImportPayroll(request, *args, **kwargs):
    message = ""

    #Making an id to make payroll is like 2025-01
    idTracker = 0
    year = datetime.now().year
    #Mapping fields from clockify to fields our models use
    fieldMap = {
        "Project": "ActivityList",
        #"Department": "dept",
        "User": "employee",
        "Start Date": "beg_date",
        "End Date": "end_date",
        #"Billable Rate (USD)": "billableRate",
        "Billable Amount (USD)": "pay_amount",
        "Duration (decimal)": "hours"
    }

    #Creating a list of activities that are tracked by employee so we can grab fund from employee not from activity
    #employeeTrackedActivities = ["AD-ADMIN", "AD-ADMIN out", "AD-COMP", "AD-COMP out", "AD-HOLIDAY", "AD-HOLIDAY out", "AD-MAC", "AD-SICK", "AD-SICK out", "AD-VAC", "AD-VAC out"]
    """activityFundMap = {
        "AD-ADMIN": "gen_pay_fund",
        "AD-ADMIN out": "gen_pay_fund",
        "AD-COMP": "comp_pay_fund",
        "AD-COMP out": "comp_pay_fund",
        "AD-HOLIDAY": "holiday_pay_fund",
        "AD-HOLIDAY out": "holiday_pay_fund",
        "AD-SICK": "sick_pay_fund",
        "AD-SICK out": "sick_pay_fund",
        "AD-VAC": "vac_pay_fund",
        "AD-VAC out": "vac_pay_fund",
        "AD-MAC": "mac_pay_fund"
    }"""

    activityFundMap = [
        "AD-COMP",
        "AD-COMP out",
        "AD-HOLIDAY",
        "AD-HOLIDAY out",
        "AD-SICK",
        "AD-SICK out",
        "AD-VAC",
        "AD-VAC out",
        "AD-MAC"
    ]

    adminCodeMap = [
        "AD-ADMIN",
        "AD-ADMIN out",
    ]


    if request.method == 'POST':
        form = FileInput(request.POST, request.FILES)
        if form.is_valid():
            selectedFile = form.cleaned_data['file']
            file = pd.read_csv(selectedFile)
            file.dropna(how='all', inplace=True)
            columns = file.columns
            row = file.iloc[0]
            data = []
            
            model = apps.get_model('WCHDApp', 'Payroll')
            fields = model._meta.get_fields()
            neededFields = []
            #Exclude autocreated fields like id
            for field in fields:
                if not field.auto_created:
                    neededFields.append(field.name)

            columns = list(columns)
            
            #Creating a list of indices that we want from columns
            neededIndexes = []
            for i in range(len(columns)):
                if columns[i] in fieldMap:
                    #updated columns[i] to the name we need for the model
                    columns[i] = fieldMap[columns[i]]
                    neededIndexes.append(i)

            #Creating a list of dictionaries for each row with the values we need
            for i in range(len(file)):
                dict = {}
                row = file.iloc[i]

                for j in neededIndexes:
                    column = columns[j]
                    if column == "beg_date" or column == "end_date":
                        payPeriodModel = apps.get_model('WCHDApp', 'PayPeriod')
                        periods = payPeriodModel.objects.all()
                        date = row[j]
                        newDate = datetime.strptime(date, "%m/%d/%Y").date()
                        dict[column] = newDate
                        for period in periods:
                            if period.periodStart <= newDate <= period.periodEnd:
                                dict['payperiod'] = period
                    else:
                        dict[column] = row[j]

                #dict['payroll_id'] = str(year)+"-"+str(idTracker)
                idTracker += 1
                data.append(dict)

            lookUpFields = []
            fks = []
            for field in fields:
                #Logic for foreign keys
                if field.is_relation:
                    fks.append(field.name)
                else:
                    lookUpFields.append(field)
            
            for line in data:
                for key in line:
                    if type(line[key]) == np.int64:
                        line[key] = int(line[key])
                    if key in fks:
                        #Linking objects with the fields we have
                        parentModel = apps.get_model('WCHDApp', key)
                        if key == "employee":
                            names = line[key].split(" ")
                            line[key] = parentModel.objects.get(first_name=names[0], surname=names[1])
                        elif key == "ActivityList":
                            line[key] = parentModel.objects.get(program=line[key])
                        elif key == "dept":
                            line[key] = parentModel.objects.get(dept_name=line[key])
                #print(line)
                activity = line['ActivityList']
                activityName = activity.program
                if activityName in activityFundMap:
                    employee = line['employee']
                    fund = employee.specialFund
                elif activityName in adminCodeMap:
                    employee = line['employee']
                    fund = employee.adminPayFund
                else:
                    fund = activity.fund
                #This is getting the total from clockify which ALyssa said isnt right all the time
                """
                rate = line['pay_amount']
                hours = line['hours']
                amount = rate * hours
                """

                payRate = float(line['employee'].pay_rate)
                hours = line['hours']
                amount = payRate*hours
                #Old way of testing
                #amount = line['pay_amount']
                balance = float(fund.fund_cash_balance)
                if balance > amount:
                    balance -= amount
                    fund.fund_cash_balance = balance
                    fund.save()
                else:
                    message += f"Fund doesn't have enough money: {fund} transaction skipped"
                obj, _ = model.objects.update_or_create(
                    **line,
                    defaults = line
                )    
    else:
        form = FileInput()
    
        
    return render(request, "WCHDApp/clockifyImportPayroll.html", {"form": form, "message": message})

def calculateActivitySelect(request, *args, **kwargs):
    payrollModel = apps.get_model('WCHDApp', 'Payroll')
    payperiodGroup = request.GET.get('payperiodDropdown')
    print(payperiodGroup)
    if payperiodGroup == None:
        data = payrollModel.objects.all()
    else:
        data = payrollModel.objects.filter(payperiod=payperiodGroup)
    #for obj in data:
        #print(obj.ActivityList.fund)

    #print(data)
    fields = []
    verboseNames = []
    for field in payrollModel._meta.get_fields():
        if field.is_relation:
            #Grab related model. This is why foreign keys have to be named after the model 
            parentModel = apps.get_model('WCHDApp', field.name)

            #Get the related models primary key
            fkName = parentModel._meta.pk.name
            verboseNames.append(parentModel._meta.pk.verbose_name)
            fields.append(fkName)
        else:
            verboseNames.append(field.verbose_name)
            fields.append(field.name)

    #Making the dropdowns for selecting the fund, activity, and employee
    activityModel = apps.get_model('WCHDApp', 'ActivityList')
    activities = activityModel.objects.all()
    activityChoices = []
    for activity in activities:
        activityChoices.append((activity.ActivityList_id, activity.program))
    
    employeeModel = apps.get_model("WCHDApp", "employee")
    employees = employeeModel.objects.all()
    employeeChoices = []
    for employee in employees:
        employeeChoices.append((employee.employee_id, employee.first_name))

    fundModel = apps.get_model("WCHDApp", "fund")
    funds = fundModel.objects.all()
    fundChoices = []
    for fund in funds:
        fundChoices.append((fund.fund_id, fund.fund_name))

    payperiodModel = apps.get_model("WCHDApp", "PayPeriod")
    payperiods = payperiodModel.objects.all()
    payperiodChoices = []
    for payperiod in payperiods:
        payperiodChoices.append(payperiod.payperiod_id)
    
    context = {
        "activityChoices": activityChoices,
        "employeeChoices": employeeChoices,
        "fundChoices": fundChoices,
        "payperiodChoices": payperiodChoices,
        "data": data, 
        "fields": fields, 
        "verboseFields": verboseNames}

    return render(request, "WCHDApp/calculateActivitySelect.html", context)

def getActivities(request):
    #This is used to have an array of the activities in javascript
    activityModel = apps.get_model('WCHDApp', 'ActivityList')
    activities = activityModel.objects.all()
    activityChoices = []
    for activity in activities:
        activityChoices.append((activity.ActivityList_id, activity.program))
    
    data = {
        "activities": activityChoices
    }
    return JsonResponse(data)

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def payrollView(request, *args, **kwargs):
    payperiodModel = apps.get_model("WCHDApp", "PayPeriod")
    payperiods = payperiodModel.objects.all()
    payperiodChoices = []
    for payperiod in payperiods:
        payperiodChoices.append(payperiod.payperiod_id)

    activityModel = apps.get_model('WCHDApp', 'ActivityList')
    activities = activityModel.objects.all()
    activityChoices = []
    for activity in activities:
        activityChoices.append((activity.ActivityList_id, activity.program))
    
    employeeModel = apps.get_model("WCHDApp", "employee")
    employees = employeeModel.objects.all()
    employeeChoices = []
    for employee in employees:
        employeeChoices.append((employee.employee_id, employee.first_name))

    fundModel = apps.get_model("WCHDApp", "fund")
    funds = fundModel.objects.all()
    fundChoices = []
    for fund in funds:
        fundChoices.append((fund.fund_id, fund.fund_name))
    
    context = {
        "payperiodChoices": payperiodChoices,
        "activityChoices": activityChoices,
        "fundChoices": fundChoices,
        "employeeChoices": employeeChoices,
        }

    return render(request, "WCHDApp/payrollView.html", context)

def fundSummary(request):
    fundID = request.GET.get("fundDropdown")
    payperiodID = request.GET.get('payperiodDropdown')
    if fundID != "EMPTY":
        fund = apps.get_model("WCHDApp", "fund")
        selectedFund = fund.objects.get(fund_id=fundID)
        fundName = selectedFund.fund_name

        payrollModel = apps.get_model('WCHDApp', 'Payroll')
        #Have to use double underscore instead of dot here for whatever reason
        filteredRows = payrollModel.objects.filter(ActivityList__fund__fund_id=fundID, payperiod__payperiod_id=payperiodID)

        totalPay = 0
        totalHours = 0
        for row in filteredRows:
            totalPay += row.pay_amount
            totalHours += row.hours

        context = {
            "specifiedField": "Fund Name",
            "specifiedValue": fundName,
            "sum": totalPay,
            "totalHours": totalHours
        }
        
        return render(request, "WCHDApp/partials/totalsOutput.html", context)
    else:
        return HttpResponse("No Fund selected", status=204)

def activitySummary(request):
    activityID = request.GET.get("activityDropdown")
    payperiodID = request.GET.get('payperiodDropdown')
    if activityID != "EMPTY":
        activity = apps.get_model("WCHDApp", "ActivityList")
        selectedActivity = activity.objects.get(ActivityList_id=activityID)
        activityName = selectedActivity.program

        payrollModel = apps.get_model('WCHDApp', 'Payroll')
        #Have to use double underscore instead of dot here for whatever reason
        filteredRows = payrollModel.objects.filter(ActivityList__ActivityList_id=activityID, payperiod__payperiod_id=payperiodID)

        totalPay = 0
        totalHours = 0
        for row in filteredRows:
            totalPay += row.pay_amount
            totalHours += row.hours

        context = {
            "specifiedField": "Activity Name",
            "specifiedValue": activityName,
            "sum": totalPay,
            "totalHours": totalHours
        }
        
        return render(request, "WCHDApp/partials/totalsOutput.html", context)
    else:
        return HttpResponse("No Activity selected", status=204)

def employeeSummary(request):
    employeeID = request.GET.get("employeeDropdown")
    payperiodID = request.GET.get('payperiodDropdown')
    if employeeID != "EMPTY":
        employee = apps.get_model("WCHDApp", "Employee")
        selectedEmployee = employee.objects.get(employee_id=employeeID)
        employeeName = selectedEmployee.first_name + " " + selectedEmployee.surname

        payrollModel = apps.get_model('WCHDApp', 'Payroll')
        #Have to use double underscore instead of dot here for whatever reason
        filteredRows = payrollModel.objects.filter(employee__employee_id=employeeID, payperiod__payperiod_id=payperiodID)

        totalPay = 0
        totalHours = 0
        for row in filteredRows:
            totalPay += row.pay_amount
            totalHours += row.hours

        activityModel = apps.get_model("WCHDApp", "ActivityList")
        activities = activityModel.objects.all()

        activitiesDict = {}
        for activity in activities:
            activityFilteredRows = payrollModel.objects.filter(ActivityList=activity, payperiod__payperiod_id=payperiodID, employee__employee_id=employeeID)
            activityPay = 0
            activityHours = 0
            for activityRow in activityFilteredRows:
                activityPay += activityRow.pay_amount
                activityHours += activityRow.hours
            
            activitiesDict[activity.program] = {"name":activity.program, "sum":activityPay, "hours":activityHours}

        context = {
            "employeeName": employeeName,
            "activitiesDict": activitiesDict,
            "sum": totalPay,
            "totalHours": totalHours
        }
        
        return render(request, "WCHDApp/partials/employeeBreakdown.html", context)
    else:
        return HttpResponse("No Employee selected", status=204)

def transactionCustomView(request):
    return render(request, "WCHDApp/transactionCustomView.html")

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def grantStats(request):
    grantModel = apps.get_model("WCHDApp", "Grant")
    grants = grantModel.objects.all()

    grantLineModel = apps.get_model("WCHDApp", "GrantLine")

    #Going to be a list of dictionaries. Each grant will have a dictionary
    grantList = []

    for grant in grants:
        grantLines = grantLineModel.objects.filter(grant = grant)

        totalBudgeted = 0
        totalSpent = 0
        totalRemaining = 0
        for grantLine in grantLines:
            totalRemaining += grantLine.line_budget_remaining
            totalSpent += grantLine.line_budget_spent
            totalBudgeted += grantLine.line_budgeted

        grantDict = {
            "grantID": grant.grant_id,
            "grantName": grant.grant_name,
            "awardAmount": grant.award_amount,
            "spent": totalSpent,
            "remaining": totalRemaining,
            "budgeted": totalBudgeted,
            "received": grant.received
        }

        grantList.append(grantDict)

    context = {
        "grantList": grantList
    }
    return render(request, "WCHDApp/grantStats.html", context)

def grantBreakdown(request):
    grantID = request.GET.get("grantID")

    grantModel = apps.get_model("WCHDApp", "Grant")
    grant = grantModel.objects.get(pk=grantID)

    grantLineModel = apps.get_model("WCHDApp", "GrantLine")
    grantLines = grantLineModel.objects.filter(grant__grant_id=grantID)

    linesList = []
    total = 0
    for line in grantLines:
        total += line.line_budgeted
        lineDict = {
            "lineName": line.line_name,
            "budgeted": line.line_budgeted,
            "remaining": line.line_budget_remaining,
            "spent": line.line_budget_spent,
            "income":line.line_total_income
        }
        linesList.append(lineDict)

    unbudgeted = grant.award_amount - total

    context = {
        "linesList": linesList,
        "unbudgeted": unbudgeted
    }

    return render(request, "WCHDApp/partials/grantBreakdownTable.html", context)

def grantLineView(request):
    grants = Grant.objects.all()
    
    context = {
        "grants": grants
    }
    return render(request, "WCHDApp/grantLineView.html", context)

def grantLineTableUpdate(request):
    message = ""
    grantID = request.GET.get("grant")
    grant = Grant.objects.get(pk=grantID)
    
    grantLines = GrantLine.objects.filter(grant=grant)

    #Getting just field names from model
    fields = GrantLine._meta.fields

    #Lists to sort fields for styling
    fieldNames = []
    decimalFields = []
    aliasNames = []

    calculatedProperties = {
        "Testing": [("fundBalanceMinus3", "Fund Balance Minus 3")],
        "Benefits": [("pers", "Public Employee Retirement System"), ("medicare", "Medicare"),("wc", "Workers Comp"), ("plar", "Paid Leave Accumulation Rate"), ("vacation", "Vacation"), ("sick", "Sick Leave"), ("holiday", "Holiday Leave"), ("total_hrly", "Total Hourly Cost"), ("percent_leave", "Percent Leave"), ("monthly_hours", "Monthly Hours"), ("board_share_hrly", "Board Share Hourly"), ("life_hourly", "Life Hourly"), ("salary", "Salary"), ("fringes", "Fringes"), ("total_comp", "Total Compensation")],
        "Payroll": [("pay_rate", "Pay Rate")],
        "Fund":[("calcRemaining", "Remaining"), ("budgeted", "Budgeted")],
        "Line": [("budgetRemaining", "Budget Remaining"), ("budgetSpent", "Budget Spent"), ("totalIncome", "Total Income")],
        "GrantLine": [("budgetRemaining", "Budget Remaining"), ("budgetSpent", "Budget Spent"), ("totalIncome", "Total Income")]
    }

    #Fields that should be accumulated
    summedFields = {
        "Fund": "fund_cash_balance", 
        "Line": "line_total_income",
        "Transaction": "amount",
    }

    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)

    if "GrantLine" in calculatedProperties:
        for property in calculatedProperties["GrantLine"]:
            #print(property)
            aliasNames.append(property[1])
            fieldNames.append(property[0])
            decimalFields.append(property[0])

    if request.method == 'POST':
        form = modelform_factory(GrantLine, exclude=["grant"])(request.POST)
        form.instance.grant = grant
    
        if form.is_valid():
            line = form.save()
            message = "Grant Line Created Successfully"
            form = modelform_factory(GrantLine, exclude=["grant"])()
        else:
            errors = form.errors
            if errors.get("line_budgeted"):
                message = errors["line_budgeted"][0]     
            if errors.get("lineType"):
                message = errors["lineType"][0]           
    else:
        form = modelform_factory(GrantLine, exclude=["grant"])()
    
    grantLines = GrantLine.objects.filter(grant=grant)

    context = {
        "fields": fieldNames, 
        "aliasNames": aliasNames, 
        "data": grantLines, 
        "decimalFields": decimalFields,
        "form": form,
        "grant": grant,
        "message": message,
        "grantAwardAmountRemaining":grant.grantAwardAmountRemaining
    }

    return render(request, "WCHDApp/partials/grantLineTableUpdate.html", context)

def testingGrantAccess(request):
    grantModel = apps.get_model("WCHDApp", "Grant")
    grants = grantModel.objects.all()
    context ={
        "grants": grants
    }

    fields = grantModel._meta.get_fields()
    print(fields)
    for field in fields:
        #if not field.auto_created or not isinstance(field, AutoField) or not field.is_relation:
        if not field.auto_created:
            print(field.name)
            if isinstance(field, ManyToManyField):
                print("Many to many field")
                print(grants[0].fund.all())
            else:
                print(getattr(grants[0], field.name))


    relatedFunds = grants[0].fund.all()
    print(relatedFunds)
    context = {
        "funds": relatedFunds
    }
    #making change
    return render(request, "WCHDApp/grantExpenseTesting.html", context)

@permission_required('WCHDApp.has_full_access', raise_exception=True)
def viewByYear(request):
    currentDate = datetime.now()
    year = currentDate.year
    years = list(range(2000, year+2))

    models = ["Line", "Fund", "Item", "Expense", "Revenue"]

    context = {
        "years": years,
        "models": models
    }

    return render(request, "WCHDApp/viewByYear.html", context)

def viewByYearPartial(request):
    message = ""
    #Any property that we define in models need to go here so our logic can include them in the table
    calculatedProperties = {
        "Testing": [("fundBalanceMinus3", "Fund Balance Minus 3")],
        "Benefits": [("pers", "Public Employee Retirement System"), ("medicare", "Medicare"),("wc", "Workers Comp"), ("plar", "Paid Leave Accumulation Rate"), ("vacation", "Vacation"), ("sick", "Sick Leave"), ("holiday", "Holiday Leave"), ("total_hrly", "Total Hourly Cost"), ("percent_leave", "Percent Leave"), ("monthly_hours", "Monthly Hours"), ("board_share_hrly", "Board Share Hourly"), ("life_hourly", "Life Hourly"), ("salary", "Salary"), ("fringes", "Fringes"), ("total_comp", "Total Compensation")],
        "Payroll": [("pay_rate", "Pay Rate")],
        "Fund":[("calcRemaining", "Remaining")]
    }

    #Requests come in as both get and post request whether it is the form being submitted or the htmx triggering the rendering
    modelName = request.GET.get('model') or request.POST.get('model')
    year = request.GET.get("year") or request.POST.get("year")
    model = apps.get_model('WCHDApp', modelName)
    if modelName == "Fund":
        values = Fund.objects.filter(fund_id__startswith=year)
        fields = Fund._meta.fields 
        if request.method == "POST":
            form = modelform_factory(model, exclude=["fund_total", "fund_budgeted", "fund_remaining"])(request.POST)
            if form.is_valid():
                fund = form.save(commit=False)
                balance = fund.fund_cash_balance
                baseID = fund.fund_id
                fund.fund_total = balance
                fund.fund_budgeted = 0
                fund.fund_remaining = balance
                currentDateTime = datetime.now()
                year = currentDateTime.year
                fullID = f"{year}-{baseID}"
                fund.fund_id = fullID

                messsage = "Fund Created"
                form.save()
        else:
            form = modelform_factory(model, exclude=["fund_total", "fund_budgeted", "fund_remaining"])(request.POST)
    if modelName == "Line":
        values = Line.objects.filter(fund__fund_id__startswith=year)
        fields = Line._meta.fields
        if request.method == 'POST':
            form = modelform_factory(Line, exclude=["line_budget_spent", "line_budget_remaining", "line_total_income"])(request.POST)
            if form.is_valid():
                line = form.save(commit=False)
                fund = line.fund
                fundID = fund.fund_id
                #Deconstructing then recontructing line id to fit county
                paritalLineID = line.line_id
                fullLineID = str(fundID)+"-"+str(paritalLineID)

                line.line_id = fullLineID

                budgeted = line.line_budgeted
                line.line_budget_spent = 0
                line.line_total_income = 0
                line.line_budget_remaining = budgeted

                line.fund = fund
                remaining = fund.fund_total - fund.fund_budgeted
                if (remaining >= budgeted):
                    fund.fund_budgeted += budgeted
                    fund.save()
                    line.save()
                    message="Line Created"
                else:
                    message="Not enough remaining balance in fund"
        else:
            form = modelform_factory(Line, exclude=["line_budget_spent", "line_budget_remaining", "line_total_income"])(request.POST)
    if modelName == "Expense":
        values = Expense.objects.filter(line__fund__fund_id__startswith=year)
        fields = Expense._meta.fields
        

    fieldNames = []
    aliasNames = []
    decimalFields = []
    for field in fields:
        if isinstance(field, DecimalField):
                decimalFields.append(field.name)
        aliasNames.append(field.verbose_name)  
        fieldNames.append(field.name)
    #Making sure properties are added like normal fields to the tables
    if modelName in calculatedProperties:
        for property in calculatedProperties[modelName]:
            #print(property)
            aliasNames.append(property[1])
            fieldNames.append(property[0])
            decimalFields.append(property[0])

    context = {"fields": fieldNames, 
               "aliasNames": aliasNames, 
               "data": values, 
               "tableName": modelName, 
               "decimalFields": decimalFields,
               "form": form,
               "year":year,
               "message": message}

    return render(request, "WCHDApp/partials/viewByYearPartial.html", context)
